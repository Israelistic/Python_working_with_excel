import openpyxl

inv_file = openpyxl.load_workbook("inventory.xlsx")
product_list = inv_file["Sheet1"]

print(product_list)

# declare a dictionaries:

products_per_supplier = {}
total_value_per_supplier = {}
product_under10_inv = {}

# print(product_list.max_row)

for product_row in range(2, product_list.max_row + 1):
    # supplier_name is from the column 4
    supplier_name = product_list.cell(product_row, 4).value
    #print(supplier_name)
    inventory = product_list.cell(product_row, 2).value
    price = product_list.cell(product_row, 3).value
    product_num = product_list.cell(product_row, 1).value
    #overwitre value in column 5
    inventory_price = product_list.cell(product_row, 5)

    # calculation number of products per supplier
    # check if the supplier name appear in dictionary if no add a new supplier if yes increment number of products
    if supplier_name in products_per_supplier:
        # assign the value of the current iteration
        current_num_of_products = products_per_supplier.get(supplier_name)
        products_per_supplier[supplier_name] = current_num_of_products + 1
    else:
        products_per_supplier[supplier_name] = 1

    # calculate total value of inventory per supplier
    if supplier_name in total_value_per_supplier:
        current_total_value = total_value_per_supplier.get(supplier_name)
        total_value_per_supplier[supplier_name] = current_total_value + inventory * price
    else:
        total_value_per_supplier[supplier_name] = inventory * price

    # Logic product with inventory less than 10
    if inventory < 10:
        product_under10_inv[int(product_num)] = int(inventory)

    # add value for total inventory price
    inventory_price.value = inventory * price




print("product inventory under 10:\n", product_under10_inv)
print("Product per supplier:\n", products_per_supplier)
print("Total value per supplier:\n", total_value_per_supplier)
inv_file.save("inventory_with_total_value.xlsx")